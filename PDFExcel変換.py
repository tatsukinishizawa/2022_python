from pdfminer3.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer3.converter import PDFPageAggregator
from pdfminer3.pdfpage import PDFPage
from pdfminer3.layout import LAParams, LTTextContainer
import tkinter.filedialog

#target_file = tkinter.filedialog.askopenfilename(filetypes="PDFファイル","*.pdf")

resourceManager = PDFResourceManager()
device = PDFPageAggregator(resourceManager, laparams=LAParams())





pdf_file_name = r"C:\Users\saitoh\PycharmProjects\transExcelFromPDF\Yageo殿向け　TS3-6205ZZNC3 TS3-6900ZZNC4.pdf"
with open(pdf_file_name, 'rb') as fp:
    interpreter = PDFPageInterpreter(resourceManager, device)
    for page in PDFPage.get_pages(fp):
        interpreter.process_page(page)
        layout = device.get_result()
        for lt in layout:
            # LTTextContainerの場合だけ標準出力
            if isinstance(lt, LTTextContainer):
                print(lt.get_text())
device.close()

#TuneupPiches

from pdfminer3.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer3.converter import PDFPageAggregator
from pdfminer3.pdfpage import PDFPage
from pdfminer3.layout import LAParams, LTTextContainer
import pandas as pd
import openpyxl
import math
from tkinter import filedialog
import os

def pdfminer_config(line_overlap, word_margin, char_margin,line_margin, detect_vertical):
    laparams = LAParams(line_overlap=line_overlap,
                        word_margin=word_margin,
                        char_margin=char_margin,
                        line_margin=line_margin,
                        detect_vertical=detect_vertical)
    resource_manager = PDFResourceManager()
    device = PDFPageAggregator(resource_manager, laparams=laparams)
    interpreter = PDFPageInterpreter(resource_manager, device)
    return (interpreter, device)

typ = [('pdfファイル','*.pdf')]

pdf_file_name = filedialog.askopenfilename(filetypes = typ, initialdir = dir)
work_file = os.path.splitext(pdf_file_name)[0] + '_work.xlsx'
excel_file_name = os.path.splitext(pdf_file_name)[0] + '.xlsx'

list1 = ['','','','','','','','']
df_x = pd.DataFrame([list1])
df_x.columns = ['page', 'word', 'x1','x2','y1','y2','width','hight']
int_page = 0
ii_index = 0

with open(pdf_file_name, 'rb') as fp:
    interpreter, device = pdfminer_config(line_overlap=0.1, word_margin=0.1,
            char_margin=0.1, line_margin=0.1, detect_vertical=False)
    for page in PDFPage.get_pages(fp):
        int_page = int_page + 1
        interpreter.process_page(page)
        layout = device.get_result()
        for lt in layout:
            # LTTextContainerの場合だけ標準出力
            if isinstance(lt, LTTextContainer):
                df_x.loc[ii_index] = [int_page,'{}'.format(lt.get_text().strip()), lt.x0 , lt.x1 ,\
                   841 - lt.y0 + (int_page - 1) * 841,841 - lt.y1  + (int_page - 1) * 841,lt.width ,lt.height ]
                ii_index = ii_index + 1

device.close()

# x1でソート
df_s_x = df_x.sort_values(['x1','y2'], ascending=[True,True])
# 縦のピッチを計算
h_min = 100
for i in range(len(df_s_x)):
    if i > 0:
        if df_s_x.iloc[i-1,2] == df_s_x.iloc[i,2]:
            h_sa = df_s_x.iloc[i,5] - df_s_x.iloc[i-1,5]
            if h_sa > 1.0 and h_min > h_sa:
                h_min = h_sa

# workファイルを書き出し
with pd.ExcelWriter(work_file) as writer:
    df_s_x.to_excel(writer, sheet_name='sheet1', index=False)
wb = openpyxl.Workbook()
ws = wb.worksheets[0]

j = 1
width_x = 0
for i in range(len(df_s_x)):
    y = df_s_x.iloc[i,5] //(math.ceil(h_min*10)/10) + 1
    c1 = ws.cell(row=int(y), column=j)
    if c1.value == None:
        c1.value = df_s_x.iloc[i,1]
    else:
        #列幅調整
        ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = (df_s_x.iloc[i,2]/5.98- width_x )
        width_x = df_s_x.iloc[i,2]/5.98
        j = j + 1
        c1 = ws.cell(row=int(y), column=j)
        c1.value = df_s_x.iloc[i,1]

wb.save(excel_file_name)